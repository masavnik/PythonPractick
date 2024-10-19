from openpyxl import load_workbook, Workbook
import os

desktop_path = os.path.join(os.path.expanduser("~"), 'OneDrive\Рабочий стол\Папа')
go_directory = os.chdir(desktop_path)

files = os.listdir(desktop_path)

excel_file = [i for i in files if i.endswith('.xlsx')]
# Создание нового рабочего листа
wb = Workbook()
ws = wb.active

# Переменная для хранения текущей строки
row = 1

# Перебор файлов и запись данных в рабочий лист
for file_name in excel_file:
    # Открытие файла
    workbook = load_workbook(file_name)
    worksheet = workbook.active

    # Копирование данных из текущего файла
    for row_data in worksheet.iter_rows(values_only=True):
        ws.append(row_data)
        row += 1

    # Увеличение пробела между данными из разных файлов
    ws.append([])
    row += 1

# Сохранение объединенного файла
wb.save('combined_file.xlsx')

# workbook = load_workbook(excel_file[0])
# worksheet = workbook.active
#
# for i in range(2, worksheet.max_row + 1):
#     try:
#         print(worksheet[f'H{i}'].value.split('х')[0])
#     except:
#         print()



