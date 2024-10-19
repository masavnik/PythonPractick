from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import requests
from bs4 import BeautifulSoup
import selenium
import pandas



workbook = load_workbook(filename='crypta.xlsx')
sheet = workbook.active


for row in range(1, sheet.max_row + 1):
    text = str(sheet[f'C{row}'].value)
    podp = str(sheet[f'E{row}'].value)
    chanel = str(sheet[f'B{row}'].value)


    if '@' in text:
        split_text = text.split()
        for i in split_text:
            if i.startswith('@'):
                # worksheet[f'A{row}'] = row
                sheet[f'G{row}'] = chanel
                sheet[f'H{row}'] = i
                sheet[f'I{row}'] = podp
                print(chanel, i, podp)

workbook.save(f'23.xlsx')