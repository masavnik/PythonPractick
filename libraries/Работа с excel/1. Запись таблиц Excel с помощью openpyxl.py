from openpyxl import Workbook

# Запись новых данных в таблицу
# workbook = load_workbook(filename="hello_world.xlsx")
# sheet = workbook.active
# sheet["C1"] = "writing ;)"
# workbook.save(filename="hello_world_append.xlsx")


filename = 'hello_world.xlsx'

# Создаем новую таблицу
workbook = Workbook()
sheet = workbook.active

# Добавляем данные таблицу
sheet['A1'] = 'hello'
sheet['B1'] = 'world'


def print_rows():
    '''
    Функция для печати данных
    '''
    for row in sheet.iter_rows(values_only=True):
        print(row)


# добавлять значения в таблицу
cell = sheet['A1']
# print(cell.value)
cell.value = 'hey'
# print(cell.value)


sheet['B10'] = 'test'

# Управление строками и столбцами
# Вставляем столбец перед существующим столбцом 1 ("A")
sheet.insert_cols(idx=1)
# print_rows()

# Вставляем 5 столбцов между 2 ("B") и 3 ("C")
sheet.insert_cols(idx=3, amount=5)
# print_rows()

# Удаляем вставленные столбцы.
sheet.delete_cols(idx=3, amount=5)
sheet.delete_cols(idx=1)
# print_rows()


# Вставляем 3 новые строки в начало.
sheet.insert_rows(idx=1, amount=3)
# print_rows()

# Удаляем первые 4 строки.
sheet.delete_rows(idx=1, amount=4)
# print_rows()


workbook.save(filename=filename)