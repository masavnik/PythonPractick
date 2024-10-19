from openpyxl import load_workbook

filename = 'hello_world.xlsx'

# Создаем новую таблицу
workbook = load_workbook(filename=filename)
sheet = workbook.active

# Вкладка
print(workbook.sheetnames)

# Выбор вкладки
sheet_one = workbook['Вкладка 1']
sheet_two = workbook['Вкладка 2']
sheet_three = workbook['Вкладка 3']

# Замена имени вкладки
products_sheet = workbook['Вкладка 1']
products_sheet.title = 'Products'

# Создание вклакди
operations_sheet = workbook.create_sheet("Operations")

# Определение положения вкладки
hr_sheet = workbook.create_sheet("HR", 0)

# Удаление вкладки
workbook.remove(operations_sheet)
workbook.remove(hr_sheet)