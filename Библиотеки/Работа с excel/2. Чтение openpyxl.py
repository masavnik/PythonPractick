from openpyxl import load_workbook

###############################################################

# Чтение таблицы
workbook = load_workbook(filename='arcticcon.xlsx')
sheet = workbook.active
print('1', sheet)  # Чтение листа

print('2', sheet['A2'].value)  # Чтение одной ячейки

###############################################################

# Чтение всех данных из таблицы
for i in range(1, 20):
    print(sheet[str(f'A{i}')].value)

###############################################################

print('3', sheet.cell(row=2, column=4).value)
# row - номер строки
# column - номер столбца

###############################################################

# Итерация
print('4', sheet['A1:C2'])
# Получить все клетки из колонки A
print('5', sheet["A"])  # Можно через for i.value
# Получить все клетки из колонок от A до B
print('6', sheet["A:B"])

print()
for row in sheet.iter_rows(min_row=1,
                           max_row=2,
                           min_col=1,
                           max_col=3):
    print(row)

print()
for column in sheet.iter_cols(min_row=1,
                              max_row=2,
                              min_col=1,
                              max_col=3):
    print(column)

print()

for value in sheet.iter_rows(min_row=1,
                             max_row=2,
                             min_col=1,
                             max_col=3,
                             values_only=True):
    print(value)

print()

###############################################################

# Работа с данными при помощи стандартных структур данных Python

for value in sheet.iter_rows(min_row=2,
                             max_row=4,
                             values_only=True):
    print(1, value)
