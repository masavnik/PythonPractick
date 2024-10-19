import json
from openpyxl import load_workbook
workbook = load_workbook(filename="reviews-sample.xlsx")
sheet = workbook.active
products = {}
# Используйте values_only потому что нам нужно получить именно значения клеток
for row in sheet.iter_rows(min_row=2,
                           min_col=4,
                           max_col=7,
                           values_only=True):
    product_id = row[0]
    product = {
        "parent": row[1],
        "title": row[2],
        "category": row[3]
    }
    products[product_id] = product
# Преобразуем в JSON, для лучшего отображения.
print(json.dumps(products, ensure_ascii=False))